
import random
import string
import socket

from tkinter import messagebox

import threading

from threading import Thread

import traceback
import customtkinter as cust

from scipy.spatial import distance as dist

from imutils.video import VideoStream
from imutils import face_utils

import imutils
import time
import dlib
import cv2
import openpyxl as excel
import pathlib
import re

lobbyname = ""
lobbycode = ""
hostlcode = ""

class INITSERVER():

    def __init__(self):

        self.flag = threading.Event()

        self.PORT = 5000    
        self.SERVER = socket.gethostbyname(socket.gethostname()) #use for multiple computers
        #self.SERVER = "" #use this for testing purposes within the same computer; caveat is server only displays results for one terminal not multiple 
        self.ADDRESS = (self.SERVER, self.PORT)
        self.FORMAT = "utf-8"

        self.clients, self.names = [], []

        self.server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        self.server.bind(self.ADDRESS)

        self.host = ""

        #self.max_threads = 30
        #self.thread_counter = 0

        #if threading.active_count() <= self.max_threads:

        self.t = Thread(target = self.startChat)
        self.t.start()




    def startChat(self):

        print("server is working on " + self.SERVER)

        self.server.listen(31)

        #self.server.settimeout(0.005) # time out after 5 ms.

        while True:
            
            try:

                self.conn, self.addr = self.server.accept()

                print ("Connection accepted")

                #self.conn.settimeout(0.005) # time out after 5 ms.

                self.msg = self.conn.recv(1024).decode(self.FORMAT)

                print (self.msg)
                print (str(self.host) + "LMAO")

                if "CL:" in self.msg:

                    if self.host != "" or self.host.isspace():

                        messagebox.showwarning("Lobby Exists!", "Host has already created a lobby!")

                    else:

                        self.host = self.conn

                        self.lobbyname = self.msg.replace("CL:", "")

                        print (self.lobbyname)

                        self.t2 = Thread(target = self.startHost)
                        self.t2.start()
                        
                        print("Thread started")

                        #self.t2.join()

                        print("Thread stopped")

                elif "RCODE:" in self.msg:
                    
                    print ("HOPE THIS WORKS")

                    self.conn.send((self.lobbycode).encode(self.FORMAT))

                    print (self.lobbycode)

                    #self.client = self.conn

                    self.t3 = Thread(target = self.startClient, args = (self.conn,))
                    self.t3.start()

                #self.clients.append(self.conn)

            except:

                print (traceback.format_exc())

    def startHost(self):

        try:

            while True:

                print ("Connection resumed")

                

                self.msg2 = self.host.recv(1024).decode(self.FORMAT)

                if "GCODE:" in self.msg2:

                    self.lobbycode = self.msg2.replace("GCODE:", "")

                    print (self.lobbycode)

                elif "CL:" in self.msg2:

                    self.lobbyname = self.msg2.replace("CL:", "")

                    print (self.lobbyname)

                elif "ON:" in self.msg2:

                    if not self.clients:

                        self.host.send(("NO CLIENTS:").encode(self.FORMAT))

                    else:

                        for x in self.clients:

                            x.send(("ON:").encode(self.FORMAT))

                elif "OFF:" in self.msg2:

                    for x in self.clients:

                            x.send(("OFF:").encode(self.FORMAT))

                else:

                    #self.host.close()
                    
                    self.host.shutdown(socket.SHUT_RDWR)

                    self.host = ""

                    #self.server.close(self.host)

                    return

        except:

            print (traceback.format_exc())
            

    def startClient(self, conn):

        try:

            while True:

                self.msg3 = conn.recv(1024).decode(self.FORMAT)

                if "CLIENT:" in self.msg3:

                    self.clients.append(conn)

                    conn.send(("CLIENT:" + self.lobbyname).encode(self.FORMAT))

                elif "NAME:" in self.msg3:

                    self.clientname = self.msg3.replace("NAME:", "")

                    print (self.clientname)

                    self.host.send(("NAME:" + self.clientname).encode(self.FORMAT))

                elif "CLOSED:" in self.msg3 or "AWAKE:" in self.msg3 or "NFD:" in self.msg3 or "SLEEPING:" in self.msg3:

                    self.host.send((self.msg3).encode(self.FORMAT))

                elif "CLEND:" in self.msg3:

                    self.host.send((self.msg3).encode(self.FORMAT))
                    conn.send(("CLEND").encode(self.FORMAT))

                    print ("CLEND successful")

                elif "AVG:" in self.msg3:

                    name = re.search(":(.+?);", self.msg3)

                    if name:

                        found = name.group(1)

                    print ("NAME IS: " + found)

                    avg_value = str(self.msg3.partition(";") [2])

                    print ("Average value is " + avg_value)

                    workbook_name = "data.xlsx"

                    path = pathlib.Path("F:\\Personal Programs\\Python\\Lobby + Tracking\\" + workbook_name) #full path of the file must be specified

                    if path.is_file():
                        
                        print ("EXCEL FILE FOUND")

                        wb = excel.load_workbook(workbook_name)
                        sheet = wb.active
                        max = sheet.max_row

                        
                        sheet.cell(row = max + 1, column = 1, value = str(found))
                        sheet.cell(row = max + 1, column = 2, value = avg_value)

                        wb.save(workbook_name)

                    else:

                        print ("EXCEL FILE NOT FOUND")

                        wb = excel.Workbook()
                        sheet = wb.active
                        #max = sheet.max_row
                        #excel_counter = 2

                        sheet["A1"] = "Client Name"
                        sheet["B1"] = "AVG_EARVALUE"

                        sheet.cell(row = 1, column = 1, value = found)
                        sheet.cell(row = 1, column = 2, value = avg_value)

                        wb.save(workbook_name)








        except:

            print (traceback.format_exc())

i = INITSERVER()


